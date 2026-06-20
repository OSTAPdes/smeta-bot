import os
import json
import time
import uuid
import base64
import asyncio
import logging
from io import BytesIO
from pathlib import Path
from contextlib import asynccontextmanager

from fastapi import FastAPI
from fastapi.responses import HTMLResponse, PlainTextResponse

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, WebAppInfo, FSInputFile

from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import anthropic

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("supply-bot")

# --- Config from environment ---
BOT_TOKEN = os.environ["BOT_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
ALLOWED_USER_IDS = {
    int(x) for x in os.environ.get("ALLOWED_USER_IDS", "").replace(" ", "").split(",") if x
}
OWNER_USER_IDS = {
    int(x) for x in os.environ.get("OWNER_USER_IDS", "").replace(" ", "").split(",") if x
}
MAX_TRIES_PER_USER = int(os.environ.get("MAX_TRIES_PER_USER", "2"))
SUPPORT_USERNAME = os.environ.get("SUPPORT_USERNAME", "").lstrip("@")
PUBLIC_URL = os.environ.get("PUBLIC_URL", "").rstrip("/")
# Когда true — бот открыт для любого пользователя Telegram, ALLOWED_USER_IDS
# не проверяется вообще. Лимит MAX_TRIES_PER_USER при этом остаётся главной
# защитой от случайного перерасхода — обязательно держи его разумным.
PUBLIC_BOT = os.environ.get("PUBLIC_BOT", "false").lower() in ("1", "true", "yes")
# Эксперимент: вместо нескольких вариантов с ценой — один максимально точный
# вариант, имя + ссылка, без цены. Меньше отвлекающих задач — больше внимания
# на само совпадение по виду.
SIMPLE_MATCH_MODE = os.environ.get("SIMPLE_MATCH_MODE", "false").lower() in ("1", "true", "yes")

CLAUDE_MODEL = os.environ.get("CLAUDE_MODEL", "claude-sonnet-4-6")  # для анализа картинки И для поиска — точность важнее экономии

BASE_DIR = Path(__file__).parent
# Если задана STORAGE_DIR (например, смонтированный Volume на Railway) — храним
# рендеры и счётчик попыток там, чтобы они переживали передеплой. Без неё —
# как раньше, в папке рядом с кодом (стирается при каждом редеплое).
STORAGE_DIR = Path(os.environ.get("STORAGE_DIR", str(BASE_DIR)))
RENDERS_DIR = STORAGE_DIR / "renders"
RENDERS_DIR.mkdir(parents=True, exist_ok=True)
USAGE_FILE = STORAGE_DIR / "usage.json"
CACHE_FILE = STORAGE_DIR / "materials_cache.json"
MAX_CACHE_PER_CATEGORY = 50
TEMPLATE = (BASE_DIR / "templates" / "app_template.html").read_text(encoding="utf-8")

claude = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

CHUNK_SIZE = 6
GRID_DIVISIONS = 10

# Эти сайты предлагаем проверить В ПЕРВУЮ ОЧЕРЕДЬ (профильные, проверенные),
# но поиск ими больше технически не ограничен — если там нет подходящего
# товара, модель имеет право поискать дальше у других надёжных украинских
# продавцов, а не выдавать что попало лишь бы заполнить позицию.
CATEGORY_DOMAINS = {
    "lighting": ["linija-svitla.ua", "svetilnikof.com.ua", "svetua.com.ua", "lampa.od.ua", "citylight.com.ua", "lustralux.com.ua"],
    "wood_decor": ["kronospan.com", "egger.com", "kronas.com.ua", "agtplus.ua", "scandiwall.com.ua"],
    "tile": ["plitka.ua", "plitkashop.com.ua", "cersanit.in.ua", "leoceramika.com", "topovi.com.ua", "stone.kiev.ua", "supers.com.ua"],
    "laminate": ["my-floor.com.ua", "parketiko.com.ua", "laminat-parketdoska.com.ua"],
    "paint": ["ncscolour.com.ua", "tikkurila-shop.com.ua", "colorstudio.com.ua"],
    "furniture": ["mebelok.com", "klen.ua", "ddn.ua", "dobralavka.ua", "taburetka.ua"],
    "quartz_marble": ["topovi.com.ua", "stone.kiev.ua", "supers.com.ua", "kitstone.kiev.ua"],
}

# А эти — технически заблокированы на уровне инструмента поиска, без исключений.
BLOCKED_DOMAINS = [
    "prom.ua", "rozetka.com.ua", "olx.ua",
    "ozon.ru", "wildberries.ru", "yandex.ru", "ya.ru", "avito.ru",
]

SUPPORT_LINE = f"\n\nПоддержка: https://t.me/{SUPPORT_USERNAME}" if SUPPORT_USERNAME else ""


def is_allowed(user_id: int) -> bool:
    if PUBLIC_BOT:
        return True
    return user_id in ALLOWED_USER_IDS or user_id in OWNER_USER_IDS


def load_usage() -> dict:
    if USAGE_FILE.exists():
        try:
            return json.loads(USAGE_FILE.read_text())
        except Exception:
            return {}
    return {}


def save_usage(data: dict):
    USAGE_FILE.write_text(json.dumps(data))


def load_cache() -> dict:
    """Общая база уже найденных материалов — одна на всех пользователей,
    растёт со временем, чем больше используется бот."""
    if CACHE_FILE.exists():
        try:
            return json.loads(CACHE_FILE.read_text())
        except Exception:
            return {}
    return {}


def append_to_cache(new_entries: list):
    """new_entries: [(category, entry_dict), ...]. Один синхронный вызов в
    конце обработки рендера — без гонок, потому что вызывается уже ПОСЛЕ
    того, как все параллельные категории закончили работу."""
    if not new_entries:
        return
    cache = load_cache()
    for category, entry in new_entries:
        bucket = cache.setdefault(category, [])
        bucket.append(entry)
        if len(bucket) > MAX_CACHE_PER_CATEGORY:
            cache[category] = bucket[-MAX_CACHE_PER_CATEGORY:]
    CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False))


_usage_lock = asyncio.Lock()


async def try_consume_quota(user_id: int) -> bool:
    """True — можно обрабатывать. Владельцам лимит не считаем.
    Проверка и запись объединены под блокировкой, чтобы два почти
    одновременных запроса от одного человека не проскочили оба разом.
    ВАЖНО: счётчик хранится в файле на диске сервиса — без подключённого
    Volume (см. STORAGE_DIR) он обнуляется при каждом передеплое. Это
    дружелюбное ограничение, а не железная защита — основной барьер от
    перерасхода всё равно держи в лимите трат на console.anthropic.com."""
    if user_id in OWNER_USER_IDS:
        return True
    async with _usage_lock:
        usage = load_usage()
        count = usage.get(str(user_id), 0)
        if count >= MAX_TRIES_PER_USER:
            return False
        usage[str(user_id)] = count + 1
        save_usage(usage)
        return True


def resize_for_claude(raw_bytes: bytes, max_side: int = 1150) -> bytes:
    img = Image.open(BytesIO(raw_bytes)).convert("RGB")
    w, h = img.size
    if max(w, h) > max_side:
        scale = max_side / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    out = BytesIO()
    img.save(out, format="JPEG", quality=85, optimize=True)
    return out.getvalue()


def build_grid_overlay(clean_jpeg_bytes: bytes):
    img = Image.open(BytesIO(clean_jpeg_bytes)).convert("RGB")
    w, h = img.size
    cell = max(w, h) / GRID_DIVISIONS
    cols = min(18, max(1, round(w / cell)))
    rows = min(30, max(1, round(h / cell)))
    col_w, row_h = w / cols, h / rows

    overlay = img.copy()
    draw = ImageDraw.Draw(overlay, "RGBA")
    line_color = (255, 60, 60, 200)
    for c in range(1, cols):
        x = c * col_w
        draw.line([(x, 0), (x, h)], fill=line_color, width=2)
    for r in range(1, rows):
        y = r * row_h
        draw.line([(0, y), (w, y)], fill=line_color, width=2)

    font = ImageFont.load_default()
    for c in range(cols):
        for r in range(rows):
            label = f"{chr(65 + c)}{r + 1}"
            x, y = c * col_w + 3, r * row_h + 2
            draw.rectangle([x - 1, y - 1, x + len(label) * 6 + 1, y + 10], fill=(0, 0, 0, 170))
            draw.text((x, y), label, fill=(255, 230, 60, 255), font=font)

    out = BytesIO()
    overlay.save(out, format="JPEG", quality=85)
    return base64.b64encode(out.getvalue()).decode(), cols, rows


def cell_to_pct(cell, cols: int, rows: int):
    try:
        cell = str(cell).strip().upper()
        col_idx = max(0, min(cols - 1, ord(cell[0]) - 65))
        row_idx = max(0, min(rows - 1, int(cell[1:]) - 1))
        return round((col_idx + 0.5) / cols * 100, 1), round((row_idx + 0.5) / rows * 100, 1)
    except Exception:
        return 50.0, 50.0


def extract_json(text: str):
    text = text.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:]
    start_candidates = [i for i in [text.find("["), text.find("{")] if i != -1]
    if not start_candidates:
        raise ValueError("No JSON found in response")
    start = min(start_candidates)
    end = max(text.rfind("]"), text.rfind("}"))
    return json.loads(text[start:end + 1])


ANALYSIS_PROMPT_TEMPLATE = """Ты — ассистент архитектора-дизайнера с очень внимательным глазом на детали.

На картинке наложена сетка: красные линии делят её на {cols} столбцов (A-{last_col}) и {rows} строк (1-{rows}), подпись каждой ячейки — жёлтым текстом в её левом верхнем углу.

ВАЖНО — что искать: только крупные архитектурные и мебельные позиции — стены, потолок, пол, диван, шкаф, кровать и прочая крупная мебель, люстры, бра и другие светильники, плитка, деревянная отделка/панели.

НЕ включай мелкие предметы и личные вещи: полотенца, столовые приборы и посуду, одежду, мелкие аксессуары, лампочки (сами лампы внутри светильника — не нужны, а вот сам светильник нужен), книги, мелкий декор на столах.

Шаг 1. Пройдись по изображению зона за зоной: потолок, стены, пол, затем крупная мебель, затем освещение (только сами светильники).

Шаг 2. Для каждой обнаруженной позиции (6-10 штук) укажи:
- id: порядковый номер начиная с 1
- title: короткое название на русском
- eyebrow: категория одним словом ("Отделка", "Мебель", "Освещение")
- desc: 1-2 предложения, только визуально подтверждённые признаки (цвет, тип поверхности, материал). Если технологию нельзя определить точно — опиши нейтрально.
- cell: код ячейки сетки (например "C4"), где находится ЦЕНТР именно этого предмета — сверь с жёлтой подписью в этой ячейке, что она попадает на сам предмет, а не на пол под ним, стену за ним или соседний объект
- unit: "м²" для отделки стен/потолка/пола, "шт." для мебели/светильников
- tiered: true для отделки деревом/ДСП или плиткой/камнем (нужен подбор: инженерный материал + натуральный аналог); false для остального
- color_match: true, если это однотонная окрашенная поверхность, для которой имеет смысл подбирать цвет по вееру NCS; иначе false
- search_category: одно из ровно этих значений — "lighting" (люстры, бра, светильники), "wood_decor" (деревянная отделка, ДСП-панели, фасады из дерева/ДСП), "tile" (керамическая плитка/керамогранит на полу/стенах), "laminate" (ламинат, паркетная доска), "paint" (просто окрашенная поверхность), "furniture" (диваны, шкафы, кровати, столы, кресла), "quartz_marble" (столешницы, подоконники, облицовка из камня/кварцевого агломерата), "other" (если не подходит ни одно)

Ответь СТРОГО в виде JSON-массива объектов с этими полями. Никакого текста до или после JSON, никакого markdown."""


VERIFY_PROMPT_TEMPLATE = """Та же картинка с сеткой ({cols} столбцов A-{last_col}, {rows} строк) и черновой список позиций:

{items_json}

Сверь КАЖДУЮ позицию с картинкой:
1. cell — жёлтая подпись этой ячейки действительно находится на названном предмете (title), а не на полу/стене под ним и не на соседнем объекте? Если нет — укажи правильную ячейку.
2. desc/title — соответствуют видимому? Перепиши нейтральнее, если материал назван слишком конкретно без визуального подтверждения.

Верни ИСПРАВЛЕННЫЙ список из {count} позиций в ТОМ ЖЕ формате (id, title, eyebrow, desc, cell, unit, tiered, color_match, search_category). Только JSON, без текста и markdown."""

# Цены Claude Sonnet 4.6: $3 / млн входных токенов, $15 / млн выходных, $0.01 за поиск.
PRICE_INPUT_PER_M = 3.0
PRICE_OUTPUT_PER_M = 15.0
PRICE_PER_SEARCH = 0.01


def estimate_call_cost(resp) -> tuple:
    """Реальная цена одного запроса к Claude, в долларах — из фактических
    токенов и количества поисков в самом ответе, а не из прикидок."""
    try:
        usage = resp.usage
        input_tokens = getattr(usage, "input_tokens", 0) or 0
        output_tokens = getattr(usage, "output_tokens", 0) or 0
        num_searches = sum(
            1 for b in resp.content
            if getattr(b, "type", "") == "server_tool_use" and getattr(b, "name", "") == "web_search"
        )
        cost = input_tokens * PRICE_INPUT_PER_M / 1_000_000 + output_tokens * PRICE_OUTPUT_PER_M / 1_000_000 + num_searches * PRICE_PER_SEARCH
        return round(cost, 4), num_searches
    except Exception:
        log.warning("Could not estimate call cost", exc_info=True)
        return 0.0, 0


def analyze_render(overlay_b64: str, media_type: str, cols: int, rows: int) -> list:
    prompt = ANALYSIS_PROMPT_TEMPLATE.format(cols=cols, last_col=chr(64 + cols), rows=rows)
    t0 = time.monotonic()
    msg = claude.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=2000,
        timeout=90.0,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": overlay_b64}},
                {"type": "text", "text": prompt},
            ],
        }],
    )
    log.info("analyze_render done in %.1fs", time.monotonic() - t0)
    cost, _ = estimate_call_cost(msg)
    log.info("analyze_render cost: $%.4f", cost)
    text = "".join(b.text for b in msg.content if b.type == "text")
    items = extract_json(text)
    for it in items:
        it["x"], it["y"] = cell_to_pct(it.get("cell"), cols, rows)
    return items, cost


def verify_items(overlay_b64: str, media_type: str, items: list, cols: int, rows: int) -> list:
    draft = [{k: v for k, v in it.items() if k not in ("x", "y", "data")} for it in items]
    prompt = VERIFY_PROMPT_TEMPLATE.format(
        cols=cols, last_col=chr(64 + cols), rows=rows,
        items_json=json.dumps(draft, ensure_ascii=False), count=len(draft),
    )
    t0 = time.monotonic()
    msg = claude.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=2000,
        timeout=90.0,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": overlay_b64}},
                {"type": "text", "text": prompt},
            ],
        }],
    )
    log.info("verify_items done in %.1fs", time.monotonic() - t0)
    cost, _ = estimate_call_cost(msg)
    log.info("verify_items cost: $%.4f", cost)
    text = "".join(b.text for b in msg.content if b.type == "text")
    try:
        fixed = extract_json(text)
    except Exception:
        log.warning("Verification pass failed to parse, using draft items")
        fixed = draft
    for it in fixed:
        it["x"], it["y"] = cell_to_pct(it.get("cell"), cols, rows)
    return fixed, cost


def build_batch_research_prompt(category: str, chunk: list, cached_entries: list) -> str:
    lines = []
    for it in chunk:
        text_blob = f"{it.get('title','')} {it.get('desc','')}".lower()
        extra = ""
        if "рейк" in text_blob or "рейч" in text_blob:
            extra = (" Это, скорее всего, рейчатая деревянная панель — по умолчанию считай её изделием "
                     "на заказ у столярной мастерской, и предлагай готовый заводской аналог только если "
                     "он реально близко совпадает по виду.")
        if it.get("color_match"):
            extra += (" Это однотонная окрашенная поверхность. Зайди на ncscolour.com.ua (официальный "
                      "представитель веера NCS Index 2050 в Украине) и подбери ближайший цвет именно по "
                      "этому каталогу — формат кода должен быть стандартным NCS, например 'NCS S 1502-Y' "
                      "(буква S, 4 цифры — чёрная/цветная составляющая, затем буква оттенка). Укажи код в "
                      "поле ncs_estimate. Также через colorstudio.com.ua или tikkurila-shop.com.ua укажи, "
                      "где в Украине можно заколеровать краску в этот код.")
        match_type = (
            "ТОЛЬКО ОДИН максимально точный вариант — конкретное название и одна рабочая ссылка на товар. "
            "Цену, наличие и второй вариант НЕ указывай, даже если знаешь — сфокусируйся целиком на точности "
            "самого совпадения по виду." if SIMPLE_MATCH_MODE
            else ('двухуровневый — сначала инженерный материал/декор, затем натуральный аналог' if it.get('tiered')
                  else 'обычный — 1 максимально точный вариант, второй только если тоже очень похож')
        )
        lines.append(
            f"- id {it['id']}: {it.get('title')} ({it.get('eyebrow')}). Описание: {it.get('desc')}. "
            f"Тип подбора: {match_type}."
            f"{extra}"
        )
    items_block = "\n".join(lines)

    if cached_entries:
        cache_lines = "\n".join(
            f"  [{e['cache_id']}] {e['title']} — {e['desc']}" for e in cached_entries
        )
        cache_block = f"""
УЖЕ НАЙДЕННЫЕ РАНЕЕ ВАРИАНТЫ В ЭТОЙ КАТЕГОРИИ (из прошлых рендеров, бот их запомнил):
{cache_lines}

Для каждой новой позиции СНАЧАЛА проверь, нет ли среди них действительно похожего по цвету/фактуре/материалу
варианта. Если есть — НЕ ищи заново, просто укажи "cache_id" этого варианта вместо полей tiers/ncs_estimate.
Используй кэш только при настоящем сходстве, не просто потому что категория совпадает — лучше новый честный
поиск, чем переиспользованный, но непохожий результат."""
    else:
        cache_block = ""

    domains = CATEGORY_DOMAINS.get(category)
    if domains:
        domain_rule = (f"Сначала проверь эти профильные сайты (они для этой категории обычно самые точные "
                        f"и качественные): {', '.join(domains)}. Если на них не нашлось ничего реально похожего — "
                        f"поищи дальше на других надёжных магазинах той же тематики, украинских или зарубежных.")
    else:
        domain_rule = "Ищи на любых надёжных специализированных магазинах — украинских или зарубежных."

    simple_mode_rule = (
        "\n- РЕЖИМ ТОЧНОГО ПОДБОРА: для каждой позиции — ровно один товар, никаких альтернатив и уровней. "
        "Не указывай цену и наличие, даже примерно. Всё внимание — на то, чтобы это был ДЕЙСТВИТЕЛЬНО тот "
        "же материал/предмет по цвету, фактуре, форме и размеру, а не просто что-то из той же категории."
        if SIMPLE_MATCH_MODE else ""
    )

    return f"""Ты помогаешь архитектору найти, где купить материалы и предметы для этих позиций:

{items_block}
{cache_block}

ОБЩИЕ ПРАВИЛА:
- {domain_rule}
- Сайты могут быть из любой страны (Украина, Польша, Германия и т.д.) — главное, чтобы товар реально продавался и был похож на описание. Если для архитектора в Украине у зарубежного магазина нет прямой доставки — это нормально, он сам разберётся с логистикой, важно само совпадение по виду.
- НИКОГДА не используй российские сайты (.ru, ya.ru, ozon.ru, wildberries и подобные) и НИКОГДА не используй общие маркетплейсы (prom.ua, rozetka.com.ua, OLX и подобные) — они уже технически заблокированы, но не предлагай их и в рассуждениях.
- ТОЧНОСТЬ ВАЖНЕЕ КОЛИЧЕСТВА И ВАЖНЕЕ ЗАПОЛНЕННОСТИ. Предлагай товар, только если он РЕАЛЬНО похож на описание по цвету, фактуре, форме и материалу. Если уверенно похожего варианта только один — верни один. Максимум 2 варианта на уровень, и только если оба действительно близкие.{simple_mode_rule}
- Если после честного поиска ничего достаточно похожего не нашлось — оставь tiers пустым ("tiers": []). Это нормальный и ОЖИДАЕМЫЙ результат для редких/нестандартных позиций. Лучше честно ничего, чем случайный товар, который на самом деле не похож — архитектор должен доверять каждой ссылке, которую ты дал.
- В поле "url" — ссылка ДОЛЖНА вести на страницу КОНКРЕТНОГО найденного товара (карточка товара с фото, ценой и кнопкой купить), а не на раздел каталога и не на главную сайта.
- Цены и поставщиков бери ТОЛЬКО из реальных результатов поиска.

Ответь СТРОГО в виде JSON-объекта (без markdown, без пояснений):
{{"results": [
  {{"id": id_позиции, "cache_id": "id_из_кэша_или_null", "tiers": [{{"name": "...", "options": [{{"name":"...","supplier":"...","price_label":"...","price_uah": число_или_null,"avail":"...","url":"..."}}]}}], "ncs_estimate": "код NCS или null"}}
]}}
Если используешь cache_id — поля tiers и ncs_estimate можно оставить пустыми, они подставятся из кэша автоматически.

Включи объект для каждой из {len(chunk)} позиций выше."""


def research_batch(category: str, chunk: list, cached_entries: list) -> tuple:
    """Возвращает (data_by_id, new_cache_entries, cost_usd). new_cache_entries —
    список (category, entry) для позиций, которые искали заново (не из кэша) —
    их нужно сохранить в базу после того, как все категории закончат работу."""
    cache_by_id = {e["cache_id"]: e for e in cached_entries}
    prompt = build_batch_research_prompt(category, chunk, cached_entries)
    max_uses = min(6, max(2, len(chunk) + 1))
    tool_def = {
        "type": "web_search_20250305",
        "name": "web_search",
        "max_uses": max_uses,
        "blocked_domains": BLOCKED_DOMAINS,
    }
    t0 = time.monotonic()
    resp = claude.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=3000,
        timeout=120.0,
        tools=[tool_def],
        messages=[{"role": "user", "content": prompt}],
    )
    log.info("research_batch[%s, %d items, %d cached] done in %.1fs",
              category, len(chunk), len(cached_entries), time.monotonic() - t0)
    cost, num_searches = estimate_call_cost(resp)
    log.info("research_batch[%s] cost: $%.4f (%d searches)", category, cost, num_searches)
    text = "".join(b.text for b in resp.content if b.type == "text")
    try:
        parsed = extract_json(text)
        results = parsed.get("results", []) if isinstance(parsed, dict) else []
    except Exception:
        log.warning("Batch research parse failed for chunk %s", [it["id"] for it in chunk])
        results = []
    by_id = {r.get("id"): r for r in results if isinstance(r, dict)}

    out = {}
    new_cache_entries = []
    for it in chunk:
        r = by_id.get(it["id"])
        cache_hit = cache_by_id.get(r.get("cache_id")) if r else None
        if cache_hit:
            out[it["id"]] = {"tiers": cache_hit["tiers"], "ncs_estimate": cache_hit.get("ncs_estimate")}
        elif r:
            data = {"tiers": r.get("tiers", []), "ncs_estimate": r.get("ncs_estimate")}
            out[it["id"]] = data
            if data["tiers"]:
                entry = {
                    "cache_id": uuid.uuid4().hex[:8],
                    "title": it.get("title", ""),
                    "desc": it.get("desc", ""),
                    "tiers": data["tiers"],
                    "ncs_estimate": data["ncs_estimate"],
                }
                new_cache_entries.append((category, entry))
        else:
            out[it["id"]] = {"tiers": [], "ncs_estimate": None}
    return out, new_cache_entries, cost


def render_app_page(render_id: str, image_b64: str, items: list):
    html = TEMPLATE.replace("__IMG_B64__", image_b64).replace(
        "__ITEMS_JSON__", json.dumps(items, ensure_ascii=False)
    )
    (RENDERS_DIR / f"{render_id}.html").write_text(html, encoding="utf-8")


def meta_path(render_id: str) -> Path:
    return RENDERS_DIR / f"{render_id}_meta.json"


def save_meta(render_id: str, meta: dict):
    meta_path(render_id).write_text(json.dumps(meta, ensure_ascii=False))


def load_meta(render_id: str):
    p = meta_path(render_id)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text())
    except Exception:
        return None


# {user_id: render_id} — кто сейчас должен прислать площади для сметы (этап 3).
# Хранится в памяти процесса, сбрасывается при перезапуске — для демо-режима ок.
pending_smeta = {}
# {user_id: render_id} — на какой рендер ссылаются кнопки меню "Поставщики"/"Смета".
last_render_by_user = {}


def build_smeta_xlsx(items: list, quantities: dict, render_id: str) -> Path:
    by_id = {it["id"]: it for it in items}
    wb = Workbook()
    sh = wb.active
    sh.title = "Смета"
    headers = ["№", "Позиция", "Категория", "Ед.", "Кол-во", "Цена за ед., грн", "Сумма, грн", "Поставщик", "Ссылка"]
    sh.append(headers)
    for cell in sh[1]:
        cell.font = Font(bold=True, name="Arial")
        cell.fill = PatternFill("solid", start_color="DDDDDD")

    row = 2
    for item_id, qty in quantities.items():
        it = by_id.get(item_id)
        if not it or qty <= 0:
            continue
        tiers = (it.get("data") or {}).get("tiers") or []
        opt = tiers[0]["options"][0] if tiers and tiers[0].get("options") else None
        price = opt.get("price_uah") if opt else None
        name = (opt.get("name") if opt else None) or it.get("title")
        supplier = opt.get("supplier") if opt else "цена не найдена, уточнить вручную"
        url = opt.get("url") if opt else ""

        sh.cell(row=row, column=1, value=row - 1)
        sh.cell(row=row, column=2, value=name)
        sh.cell(row=row, column=3, value=it.get("eyebrow"))
        sh.cell(row=row, column=4, value=it.get("unit"))
        sh.cell(row=row, column=5, value=qty)
        price_cell = sh.cell(row=row, column=6, value=price if price is not None else 0)
        price_cell.number_format = '#,##0 "грн"'
        sum_cell = sh.cell(row=row, column=7, value=f"=E{row}*F{row}")
        sum_cell.number_format = '#,##0 "грн"'
        sh.cell(row=row, column=8, value=supplier)
        sh.cell(row=row, column=9, value=url)
        if price is None:
            price_cell.fill = PatternFill("solid", start_color="FFFF00")
        row += 1

    last_row = row - 1
    total_row = row + 1
    sh.cell(row=total_row, column=2, value="ИТОГО").font = Font(bold=True, name="Arial")
    total_cell = sh.cell(row=total_row, column=7, value=f"=SUM(G2:G{last_row})" if last_row >= 2 else 0)
    total_cell.font = Font(bold=True, name="Arial")
    total_cell.number_format = '#,##0 "грн"'

    for col, width in zip("ABCDEFGHI", [4, 34, 14, 6, 8, 16, 16, 26, 42]):
        sh.column_dimensions[col].width = width

    path = RENDERS_DIR / f"smeta_{render_id}.xlsx"
    wb.save(str(path))
    return path


def parse_quantities(text: str) -> dict:
    out = {}
    for line in text.strip().splitlines():
        parts = line.replace(",", ".").split()
        if len(parts) >= 2:
            try:
                out[int(parts[0])] = float(parts[1])
            except ValueError:
                continue
    return out


# --- Telegram handlers ---

MENU_KB = InlineKeyboardMarkup(inline_keyboard=[
    [InlineKeyboardButton(text="🚀 Старт", callback_data="menu:start"),
     InlineKeyboardButton(text="🧩 Распознать", callback_data="menu:recognize")],
    [InlineKeyboardButton(text="🛒 Поставщики", callback_data="menu:suppliers"),
     InlineKeyboardButton(text="📐 Смета", callback_data="menu:smeta")],
])


async def send_menu(message: Message):
    await message.answer(
        "✨ <b>OSTAP — подбор материалов с рендера</b>\n\n"
        "🧩 <b>Распознать</b> — пришли фото, получи точный список материалов, мебели и освещения\n"
        "🛒 <b>Поставщики</b> — найду, где купить, и цены по последнему рендеру (~$5)\n"
        "📐 <b>Смета</b> — посчитаю стоимость по введённым площадям (~$5)\n\n"
        "Выбери действие:",
        reply_markup=MENU_KB,
        parse_mode="HTML",
    )


@dp.message(Command("start", "menu"))
async def cmd_start(message: Message):
    if not is_allowed(message.from_user.id):
        await message.answer(
            "Доступ к боту пока ограничен. Если тебе нужен доступ — обратись к администратору.\n"
            f"Твой Telegram ID: {message.from_user.id}{SUPPORT_LINE}"
        )
        return
    await send_menu(message)


@dp.callback_query(F.data == "menu:start")
async def menu_start(callback: CallbackQuery):
    await callback.answer()
    await send_menu(callback.message)


@dp.callback_query(F.data == "menu:recognize")
async def menu_recognize(callback: CallbackQuery):
    await callback.answer()
    await callback.message.answer(
        f"📸 Пришли фото рендера интерьера или фасада — разберу материалы, мебель и освещение.{SUPPORT_LINE}"
    )


@dp.callback_query(F.data == "menu:suppliers")
async def menu_suppliers(callback: CallbackQuery):
    user_id = callback.from_user.id
    if not is_allowed(user_id):
        await callback.answer("Доступ ограничен", show_alert=True)
        return
    render_id = last_render_by_user.get(user_id)
    await callback.answer()
    if not render_id:
        await callback.message.answer(f"Сначала пришли фото рендера — это Этап 1.{SUPPORT_LINE}")
        return
    await trigger_stage2(callback.message, user_id, render_id)


@dp.callback_query(F.data == "menu:smeta")
async def menu_smeta(callback: CallbackQuery):
    user_id = callback.from_user.id
    if not is_allowed(user_id):
        await callback.answer("Доступ ограничен", show_alert=True)
        return
    render_id = last_render_by_user.get(user_id)
    await callback.answer()
    if not render_id:
        await callback.message.answer(f"Сначала пройди Этап 1 и Этап 2 для рендера.{SUPPORT_LINE}")
        return
    await trigger_stage3(callback.message, user_id, render_id)


@dp.message(F.photo)
async def handle_photo(message: Message):
    user_id = message.from_user.id
    if not is_allowed(user_id):
        await message.answer(
            f"Доступ к боту пока ограничен. Твой Telegram ID: {user_id} — "
            f"передай его администратору, чтобы получить доступ.{SUPPORT_LINE}"
        )
        return

    if message.media_group_id:
        await message.answer(
            f"Пришли, пожалуйста, одну картинку отдельным сообщением (не альбомом из нескольких фото) — "
            f"каждая попытка считается строго по одному фото.{SUPPORT_LINE}"
        )
        return

    if not await try_consume_quota(user_id):
        await message.answer(
            f"Пробный лимит ({MAX_TRIES_PER_USER} разбора) на этом аккаунте исчерпан. "
            f"Если нужно больше — напиши в поддержку.{SUPPORT_LINE}"
        )
        return

    status = await message.answer("Этап 1: распознаю материалы и предметы на рендере…")
    t0 = time.monotonic()

    photo = message.photo[-1]
    file = await bot.get_file(photo.file_id)
    file_io = await bot.download_file(file.file_path)
    clean_bytes = resize_for_claude(file_io.read())
    img_b64 = base64.b64encode(clean_bytes).decode()
    overlay_b64, cols, rows = build_grid_overlay(clean_bytes)

    try:
        items, cost1 = await asyncio.to_thread(analyze_render, overlay_b64, "image/jpeg", cols, rows)
        items, cost2 = await asyncio.to_thread(verify_items, overlay_b64, "image/jpeg", items, cols, rows)
    except Exception as e:
        log.exception("analyze_render failed")
        await status.edit_text(f"Не получилось разобрать рендер: {e}{SUPPORT_LINE}")
        return
    total_cost = cost1 + cost2
    for it in items:
        it["data"] = {"tiers": [], "ncs_estimate": None}

    render_id = uuid.uuid4().hex[:10]
    render_app_page(render_id, img_b64, items)
    save_meta(render_id, {"user_id": user_id, "img_b64": img_b64, "items": items, "stage2_done": False})
    last_render_by_user[user_id] = render_id
    log.info("Render %s stage1 done in %.1fs, real cost: $%.4f", render_id, time.monotonic() - t0, total_cost)

    if not PUBLIC_URL:
        await status.edit_text("Карта готова, но PUBLIC_URL ещё не настроен — добавь его и пришли рендер ещё раз.")
        return

    names = "\n".join(f"• {it['title']}" for it in items)
    url = f"{PUBLIC_URL}/app/{render_id}"
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Открыть карту материалов", web_app=WebAppInfo(url=url))],
        [InlineKeyboardButton(text="Этап 2: найти где купить (~$5)", callback_data=f"stage2:{render_id}")],
    ])
    cost_line = f"\n\nРеальная цена этапа 1: ${total_cost:.4f}" if user_id in OWNER_USER_IDS else ""
    await status.edit_text(
        f"Этап 1 готов, найдено {len(items)} позиций:\n{names}\n\nСсылка на карту:\n{url}{cost_line}{SUPPORT_LINE}",
        reply_markup=kb,
    )


async def trigger_stage2(reply_to: Message, user_id: int, render_id: str):
    meta = load_meta(render_id)
    if not meta:
        await reply_to.answer(f"Рендер не найден (возможно, был передеплой).{SUPPORT_LINE}")
        return
    if meta.get("stage2_done"):
        url = f"{PUBLIC_URL}/app/{render_id}"
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🗺 Открыть карту материалов", web_app=WebAppInfo(url=url))],
            [InlineKeyboardButton(text="📐 Этап 3 — смета", callback_data=f"stage3:{render_id}")],
        ])
        await reply_to.answer(f"Этап 2 для этого рендера уже выполнен.{SUPPORT_LINE}", reply_markup=kb)
        return

    status = await reply_to.answer("🛒 Этап 2: ищу, где купить, и цены — это может занять пару минут…")
    t0 = time.monotonic()

    items = meta["items"]
    groups = {}
    for it in items:
        groups.setdefault(it.get("search_category", "other"), []).append(it)

    full_cache = load_cache()
    tasks = []
    for cat, group_items in groups.items():
        for i in range(0, len(group_items), CHUNK_SIZE):
            tasks.append((cat, group_items[i:i + CHUNK_SIZE]))

    async def do_task(cat, chunk):
        cached_entries = full_cache.get(cat, [])
        try:
            return await asyncio.to_thread(research_batch, cat, chunk, cached_entries)
        except Exception:
            log.exception("research_batch failed for category %s (%d items)", cat, len(chunk))
            return {it["id"]: {"tiers": [], "ncs_estimate": None} for it in chunk}, [], 0.0

    task_results = await asyncio.gather(*(do_task(c, ch) for c, ch in tasks))
    data_by_id = {}
    all_new_cache_entries = []
    total_cost = 0.0
    for cr, new_entries, cost in task_results:
        data_by_id.update(cr)
        all_new_cache_entries.extend(new_entries)
        total_cost += cost
    for it in items:
        it["data"] = data_by_id.get(it["id"], {"tiers": [], "ncs_estimate": None})
    if all_new_cache_entries:
        append_to_cache(all_new_cache_entries)

    render_app_page(render_id, meta["img_b64"], items)
    meta["items"] = items
    meta["stage2_done"] = True
    save_meta(render_id, meta)
    log.info("Render %s stage2 done in %.1fs, real cost: $%.4f", render_id, time.monotonic() - t0, total_cost)

    url = f"{PUBLIC_URL}/app/{render_id}"
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🗺 Открыть карту материалов", web_app=WebAppInfo(url=url))],
        [InlineKeyboardButton(text="📐 Этап 3 — смета", callback_data=f"stage3:{render_id}")],
    ])
    cost_line = f"\n\n💳 Реальная цена этапа 2: ${total_cost:.4f}" if user_id in OWNER_USER_IDS else ""
    await status.edit_text(f"✅ Этап 2 готов — цены и поставщики добавлены на карту.{cost_line}{SUPPORT_LINE}", reply_markup=kb)


async def trigger_stage3(reply_to: Message, user_id: int, render_id: str):
    meta = load_meta(render_id)
    if not meta:
        await reply_to.answer(f"Рендер не найден (возможно, был передеплой).{SUPPORT_LINE}")
        return
    if not meta.get("stage2_done"):
        await reply_to.answer(f"Сначала нужно выполнить Этап 2 — без цен смету не посчитать.{SUPPORT_LINE}")
        return

    pending_smeta[user_id] = render_id
    lines = "\n".join(f"{it['id']} — {it['title']} ({it['unit']})" for it in meta["items"])
    await reply_to.answer(
        "📐 Этап 3: пришли одним сообщением площади/количества по каждой позиции, "
        "по одной строке в формате \"номер количество\", например:\n\n1 25.5\n2 12\n3 4\n\n"
        f"Позиции:\n{lines}{SUPPORT_LINE}"
    )


@dp.callback_query(F.data.startswith("stage2:"))
async def handle_stage2(callback: CallbackQuery):
    user_id = callback.from_user.id
    if not is_allowed(user_id):
        await callback.answer("Доступ ограничен", show_alert=True)
        return
    render_id = callback.data.split(":", 1)[1]
    await callback.answer()
    await trigger_stage2(callback.message, user_id, render_id)


@dp.callback_query(F.data.startswith("stage3:"))
async def handle_stage3(callback: CallbackQuery):
    user_id = callback.from_user.id
    if not is_allowed(user_id):
        await callback.answer("Доступ ограничен", show_alert=True)
        return
    render_id = callback.data.split(":", 1)[1]
    await callback.answer()
    await trigger_stage3(callback.message, user_id, render_id)


@dp.message(F.text & ~F.text.startswith("/"))
async def handle_text(message: Message):
    user_id = message.from_user.id
    render_id = pending_smeta.get(user_id)
    if not render_id:
        return
    meta = load_meta(render_id)
    if not meta:
        await message.answer(f"Рендер не найден, начни заново с фото.{SUPPORT_LINE}")
        pending_smeta.pop(user_id, None)
        return

    quantities = parse_quantities(message.text)
    if not quantities:
        await message.answer(
            f"Не разобрал формат — пришли по одной строке \"номер количество\", например:\n1 25.5\n2 12{SUPPORT_LINE}"
        )
        return

    path = build_smeta_xlsx(meta["items"], quantities, render_id)
    pending_smeta.pop(user_id, None)
    await message.answer_document(
        FSInputFile(str(path)),
        caption=f"Этап 3 готов — смета на чистовые материалы.{SUPPORT_LINE}",
    )


# --- FastAPI app ---

@asynccontextmanager
async def lifespan(app: FastAPI):
    polling_task = asyncio.create_task(dp.start_polling(bot))
    log.info("Bot polling started")
    yield
    polling_task.cancel()


app = FastAPI(lifespan=lifespan)


@app.get("/")
def health():
    return PlainTextResponse("OK")


@app.get("/app/{render_id}")
def get_render(render_id: str):
    path = RENDERS_DIR / f"{render_id}.html"
    if not path.exists():
        return HTMLResponse("<h1>Не найдено</h1>", status_code=404)
    return HTMLResponse(path.read_text(encoding="utf-8"))
